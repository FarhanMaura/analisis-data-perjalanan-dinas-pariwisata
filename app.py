from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import pandas as pd
import numpy as np
import sqlite3
import os
from datetime import datetime, timedelta
import json
from ml_analysis import TourismAnalyzer
from data_processor import DataProcessor
from chart_generator import ChartGenerator
from pdf_processor import PDFProcessor
from utils import setup_logging, create_response, validate_year
from config import Config
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
import io
import base64
import matplotlib.pyplot as plt
import matplotlib
from datetime import datetime
matplotlib.use('Agg')  # Important for generating images without GUI

# Authentication imports
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import User, HotelData, TourismData
from decorators import role_required
from export_utils import (
    export_hotel_to_excel, export_hotel_to_csv, export_hotel_to_pdf,
    export_tourism_to_excel, export_tourism_to_csv, export_tourism_to_pdf
)
import random

app = Flask(__name__)
app.secret_key = Config.SECRET_KEY
app.config['UPLOAD_FOLDER'] = Config.UPLOAD_FOLDER
app.config['DATABASE'] = Config.DATABASE
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_CONTENT_LENGTH
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu'
login_manager.login_message_category = 'error'

@login_manager.user_loader
def load_user(user_id):
    return User.get_by_id(Config.DATABASE, int(user_id))

data_processor = DataProcessor(Config.DATABASE)
ml_analyzer = TourismAnalyzer(Config.DATABASE)
chart_generator = ChartGenerator(ml_analyzer)
pdf_processor = PDFProcessor()

setup_logging()

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def init_db():
    conn = sqlite3.connect(app.config['DATABASE'])
    cursor = conn.cursor()
    
    # Existing tourism_data table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tourism_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER,
            month TEXT,
            value INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Existing uploaded_files table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS uploaded_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            year INTEGER,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # New users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            email TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # New hotel_info table (stores hotel name and total rooms)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS hotel_info (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            hotel_name TEXT NOT NULL,
            total_rooms INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # New hotel_data table (daily hotel data)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS hotel_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            occupied_rooms INTEGER NOT NULL,
            guest_count INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # New tourism_site_data table (renamed from tourism_data to avoid conflict)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tourism_site_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            origin TEXT NOT NULL,
            total_visitors INTEGER NOT NULL,
            male_adult INTEGER NOT NULL,
            female_adult INTEGER NOT NULL,
            male_child INTEGER NOT NULL,
            female_child INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    conn.commit()
    conn.close()

def create_default_admin():
    """Create default admin user if not exists"""
    conn = sqlite3.connect(app.config['DATABASE'])
    cursor = conn.cursor()
    
    # Check if admin exists
    cursor.execute('SELECT id FROM users WHERE username = ?', ('admin',))
    if not cursor.fetchone():
        password_hash = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO users (username, password_hash, role, email)
            VALUES (?, ?, ?, ?)
        ''', ('admin', password_hash, 'admin', 'admin@pariwisata.go.id'))
        conn.commit()
        print("✅ Default admin user created (username: admin, password: admin123)")
    
    conn.close()

def get_db_connection():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn

def get_data_complexity_level():
    conn = get_db_connection()
    cursor = conn.execute('SELECT COUNT(DISTINCT year) as year_count FROM tourism_data')
    year_count = cursor.fetchone()['year_count']
    cursor = conn.execute('SELECT COUNT(*) as file_count FROM uploaded_files')
    file_count = cursor.fetchone()['file_count']
    conn.close()
    return max(year_count, file_count)

def process_csv_file_simple(filepath, year):
    try:
        df = pd.read_csv(filepath)
        palembang_data = None
        for index, row in df.iterrows():
            for i, cell in enumerate(row):
                if 'Palembang' in str(cell):
                    palembang_data = row
                    break
            if palembang_data is not None:
                break
        
        if palembang_data is None:
            return False, "Data Palembang tidak ditemukan"
        
        months = ['January', 'February', 'March', 'April', 'May', 'June', 
                 'July', 'August', 'September', 'October', 'November', 'December']
        monthly_data = {}
        
        for i, month in enumerate(months):
            if i + 1 < len(palembang_data):
                value = palembang_data.iloc[i + 1]
                try:
                    value_int = int(value) if pd.notna(value) else 0
                except (ValueError, TypeError):
                    value_int = 0
                monthly_data[month] = value_int
            else:
                monthly_data[month] = 0
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM tourism_data WHERE year = ?', (year,))
        for month, value in monthly_data.items():
            cursor.execute(
                'INSERT INTO tourism_data (year, month, value) VALUES (?, ?, ?)',
                (year, month, value)
            )
        conn.commit()
        conn.close()
        
        total = sum(monthly_data.values())
        return True, f"Data berhasil diproses. Total pengunjung: {total:,}"
        
    except Exception as e:
        return False, f"Error processing CSV: {str(e)}"

def analyze_data():
    conn = get_db_connection()
    query = '''
        SELECT year, month, value 
        FROM tourism_data 
        ORDER BY year, 
        CASE month
            WHEN 'January' THEN 1
            WHEN 'February' THEN 2
            WHEN 'March' THEN 3
            WHEN 'April' THEN 4
            WHEN 'May' THEN 5
            WHEN 'June' THEN 6
            WHEN 'July' THEN 7
            WHEN 'August' THEN 8
            WHEN 'September' THEN 9
            WHEN 'October' THEN 10
            WHEN 'November' THEN 11
            WHEN 'December' THEN 12
        END
    '''
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    if df.empty:
        return {
            'monthly_data': [],
            'yearly_data': [],
            'suggestions': ["Belum ada data untuk dianalisis"],
            'charts_data': {
                'monthly_labels': [],
                'monthly_values': [],
                'yearly_labels': [], 
                'yearly_values': []
            }
        }
    
    monthly_data = []
    yearly_data = []
    
    yearly_stats = df.groupby('year')['value'].sum().reset_index()
    for _, row in yearly_stats.iterrows():
        yearly_data.append({
            'year': int(row['year']),
            'total': int(row['value'])
        })
    
    months_order = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    for month in months_order:
        month_data = df[df['month'] == month]
        if not month_data.empty:
            avg_value = month_data['value'].mean()
            monthly_data.append({
                'month': month,
                'average': int(avg_value)
            })
    
    try:
        ml_analysis = ml_analyzer.get_detailed_analysis()
        suggestions = ml_analysis['suggestions'][:3]
    except Exception as e:
        suggestions = ["Sistem analisis sedang disempurnakan"]
    
    charts_data = {
        'monthly_labels': [d['month'] for d in monthly_data],
        'monthly_values': [d['average'] for d in monthly_data],
        'yearly_labels': [str(d['year']) for d in yearly_data],
        'yearly_values': [d['total'] for d in yearly_data]
    }
    
    return {
        'monthly_data': monthly_data,
        'yearly_data': yearly_data,
        'suggestions': suggestions,
        'charts_data': charts_data
    }

def generate_monthly_chart_image(df):
    """Generate monthly chart as image"""
    if df.empty:
        return None
        
    plt.figure(figsize=(12, 6))
    
    monthly_avg = df.groupby('month')['value'].mean()
    months_order = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    monthly_avg = monthly_avg.reindex(months_order)
    
    colors = []
    if not monthly_avg.empty:
        high_threshold = monthly_avg.quantile(0.70)
        low_threshold = monthly_avg.quantile(0.30)
        
        for value in monthly_avg:
            if value >= high_threshold:
                colors.append('#FF6B6B')  # Red for High
            elif value <= low_threshold:
                colors.append('#45B7D1')  # Blue for Low
            else:
                colors.append('#4ECDC4')  # Green for Medium
    
    bars = plt.bar(months_order, monthly_avg.values, color=colors, edgecolor='#2C3E50', linewidth=1)
    plt.title('Performa Bulanan dengan Kategori Musim', fontsize=14, fontweight='bold', pad=20)
    plt.xlabel('Bulan')
    plt.ylabel('Rata-rata Pengunjung')
    plt.xticks(rotation=45)
    plt.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height):,}',
                ha='center', va='bottom', fontweight='bold')
    
    plt.tight_layout()
    
    # Save to bytes
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def generate_yearly_chart_image(df):
    """Generate yearly trend chart as image"""
    if df.empty:
        return None
        
    plt.figure(figsize=(12, 6))
    
    yearly_totals = df.groupby('year')['value'].sum().reset_index()
    yearly_totals = yearly_totals.sort_values('year')
    
    plt.plot(yearly_totals['year'], yearly_totals['value'], 
             marker='o', linewidth=3, markersize=8, color='#4ECDC4')
    plt.fill_between(yearly_totals['year'], yearly_totals['value'], alpha=0.2, color='#4ECDC4')
    
    plt.title('Trend Kunjungan Wisata Tahunan Palembang', fontsize=14, fontweight='bold', pad=20)
    plt.xlabel('Tahun')
    plt.ylabel('Total Pengunjung')
    plt.grid(True, alpha=0.3)
    
    # Add value labels on points
    for i, (year, value) in enumerate(zip(yearly_totals['year'], yearly_totals['value'])):
        plt.annotate(f'{int(value):,}', (year, value), 
                    textcoords="offset points", xytext=(0,10), 
                    ha='center', fontweight='bold')
    
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def generate_seasonal_pie_chart_image(df):
    """Generate seasonal pie chart as image"""
    if df.empty:
        return None
        
    plt.figure(figsize=(10, 8))
    
    monthly_avg = df.groupby('month')['value'].mean()
    months_order = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    monthly_avg = monthly_avg.reindex(months_order)
    
    if monthly_avg.empty:
        return None
    
    total_visitors = monthly_avg.sum()
    high_threshold = monthly_avg.quantile(0.70)
    low_threshold = monthly_avg.quantile(0.30)
    
    high_season_visitors = monthly_avg[monthly_avg >= high_threshold].sum()
    low_season_visitors = monthly_avg[monthly_avg <= low_threshold].sum()
    medium_season_visitors = total_visitors - high_season_visitors - low_season_visitors
    
    if total_visitors > 0:
        high_percentage = (high_season_visitors / total_visitors) * 100
        medium_percentage = (medium_season_visitors / total_visitors) * 100
        low_percentage = (low_season_visitors / total_visitors) * 100
    else:
        high_percentage = medium_percentage = low_percentage = 0
    
    sizes = [high_percentage, medium_percentage, low_percentage]
    labels = [f'High Season\n{high_percentage:.1f}%', 
              f'Medium Season\n{medium_percentage:.1f}%', 
              f'Low Season\n{low_percentage:.1f}%']
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
    
    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90,
            textprops={'fontweight': 'bold'})
    plt.title('Distribusi Pengunjung Berdasarkan Musim', fontsize=14, fontweight='bold', pad=20)
    
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def generate_comparison_chart_image(df):
    """Generate comparison chart as image"""
    if df.empty:
        return None
        
    plt.figure(figsize=(12, 6))
    
    years = sorted(df['year'].unique())
    if len(years) < 2:
        return None
        
    recent_years = years[-2:]
    months_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    colors = ['#FF6B6B', '#4ECDC4']
    
    for i, year in enumerate(recent_years):
        year_data = df[df['year'] == year]
        monthly_data = []
        
        for month in ['January', 'February', 'March', 'April', 'May', 'June',
                     'July', 'August', 'September', 'October', 'November', 'December']:
            month_data = year_data[year_data['month'] == month]
            if not month_data.empty:
                monthly_data.append(int(month_data['value'].iloc[0]))
            else:
                monthly_data.append(0)
        
        plt.plot(months_order, monthly_data, marker='o', linewidth=2, 
                label=f'Tahun {int(year)}', color=colors[i])
    
    plt.title(f'Perbandingan Bulanan {int(recent_years[0])} vs {int(recent_years[1])}', 
              fontsize=14, fontweight='bold', pad=20)
    plt.xlabel('Bulan')
    plt.ylabel('Jumlah Pengunjung')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def _create_raw_data_sheet(worksheet, df):
    """Create raw data sheet"""
    headers = ['Tahun', 'Bulan', 'Jumlah Pengunjung']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    if not df.empty:
        for row, (_, data_row) in enumerate(df.iterrows(), 2):
            worksheet.cell(row=row, column=1, value=int(data_row['year']))
            worksheet.cell(row=row, column=2, value=data_row['month'])
            worksheet.cell(row=row, column=3, value=int(data_row['value']))
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def _create_ml_analysis_sheet(worksheet, ml_analysis):
    """Create ML analysis sheet"""
    worksheet.cell(row=1, column=1, value="Analisis Machine Learning - Pariwisata Palembang")
    worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    worksheet.cell(row=3, column=1, value="Saran & Rekomendasi:")
    worksheet.cell(row=3, column=1).font = Font(bold=True)
    
    row = 4
    if 'suggestions' in ml_analysis and ml_analysis['suggestions']:
        for suggestion in ml_analysis['suggestions']:
            worksheet.cell(row=row, column=1, value=f"• {suggestion}")
            row += 1
    else:
        worksheet.cell(row=row, column=1, value="Tidak ada saran yang tersedia")
        row += 1
    
    row += 1
    worksheet.cell(row=row, column=1, value="Pola yang Teridentifikasi:")
    worksheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    if 'patterns' in ml_analysis and ml_analysis['patterns']:
        patterns = ml_analysis['patterns']
        if 'trends' in patterns and patterns['trends']:
            for trend in patterns['trends']:
                worksheet.cell(row=row, column=1, 
                             value=f"Trend {trend['period']}: {trend['direction']} {abs(trend['growth']):.1f}%")
                row += 1
        
        if 'seasonal_distribution' in patterns:
            seasonal = patterns['seasonal_distribution']
            if 'season_percentages' in seasonal:
                worksheet.cell(row=row, column=1, value="Distribusi Musim:")
                row += 1
                for season, percentage in seasonal['season_percentages'].items():
                    worksheet.cell(row=row, column=1, 
                                 value=f"  {season}: {percentage}%")
                    row += 1
            
            # [NEW] Add Clustering Metrics Section
            if 'clustering_metrics' in seasonal:
                row += 1
                metrics = seasonal['clustering_metrics']
                worksheet.cell(row=row, column=1, value="Validasi Model Clustering:")
                worksheet.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                if 'silhouette_score' in metrics:
                    score = metrics['silhouette_score']
                    eval_text = "Good" if score > 0.5 else "Moderate" if score > 0.25 else "Weak"
                    worksheet.cell(row=row, column=1, value=f"- Silhouette Score: {score} ({eval_text})")
                    row += 1
                
                if 'method' in metrics:
                    worksheet.cell(row=row, column=1, value=f"- Metode: {metrics['method']}")
                    row += 1

def _create_charts_sheet(worksheet, df):
    """Create charts visualization sheet with embedded images"""
    worksheet.cell(row=1, column=1, value="Visualisasi Data - Grafik dan Chart")
    worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    current_row = 3
    
    try:
        # Generate and embed monthly chart
        monthly_img = generate_monthly_chart_image(df)
        if monthly_img:
            worksheet.cell(row=current_row, column=1, value="1. Grafik Rata-rata Bulanan dengan Kategori Musim")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            img = Image(monthly_img)
            img.width = 600
            img.height = 300
            worksheet.add_image(img, f'A{current_row}')
            current_row += 20  # Move down for next chart
        
        # Generate and embed yearly chart
        yearly_img = generate_yearly_chart_image(df)
        if yearly_img:
            worksheet.cell(row=current_row, column=1, value="2. Trend Kunjungan Wisata Tahunan")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            img = Image(yearly_img)
            img.width = 600
            img.height = 300
            worksheet.add_image(img, f'A{current_row}')
            current_row += 20
        
        # Generate and embed seasonal pie chart
        seasonal_img = generate_seasonal_pie_chart_image(df)
        if seasonal_img:
            worksheet.cell(row=current_row, column=1, value="3. Distribusi Pengunjung Berdasarkan Musim")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            img = Image(seasonal_img)
            img.width = 500
            img.height = 400
            worksheet.add_image(img, f'A{current_row}')
            current_row += 25
        
        # Generate and embed comparison chart
        comparison_img = generate_comparison_chart_image(df)
        if comparison_img:
            worksheet.cell(row=current_row, column=1, value="4. Perbandingan Tahun")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            img = Image(comparison_img)
            img.width = 600
            img.height = 300
            worksheet.add_image(img, f'A{current_row}')
            
    except Exception as e:
        worksheet.cell(row=current_row, column=1, value=f"Error generating charts: {str(e)}")
        current_row += 1

def _create_statistics_sheet(worksheet, df, ml_analysis):
    """Create statistics summary sheet"""
    worksheet.cell(row=1, column=1, value="Statistik Summary - Data Pariwisata")
    worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    row = 3
    
    stats = [
        ("Total Tahun Data", len(df['year'].unique()) if not df.empty else 0),
        ("Total Records Data", len(df)),
        ("Total Pengunjung", int(df['value'].sum()) if not df.empty else 0),
        ("Rata-rata Bulanan", float(df['value'].mean()) if not df.empty else 0),
        ("Pengunjung Tertinggi", int(df['value'].max()) if not df.empty else 0),
        ("Pengunjung Terendah", int(df['value'].min()) if not df.empty else 0),
        ("Tahun Terlama", int(df['year'].min()) if not df.empty else "N/A"),
        ("Tahun Terbaru", int(df['year'].max()) if not df.empty else "N/A"),
    ]
    
    worksheet.cell(row=row, column=1, value="Statistik Dasar")
    worksheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for stat_name, stat_value in stats:
        worksheet.cell(row=row, column=1, value=stat_name)
        worksheet.cell(row=row, column=2, value=stat_value)
        row += 1
    
    row += 1
    
    if 'summary' in ml_analysis:
        worksheet.cell(row=row, column=1, value="Summary Analisis ML")
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        ml_summary = ml_analysis['summary']
        ml_stats = [
            ("Total Tahun", ml_summary.get('total_years', 0)),
            ("Total Pengunjung", ml_summary.get('total_visitors', 0)),
            ("Rata-rata Bulanan", ml_summary.get('avg_monthly', 0)),
            ("Periode Data", ml_summary.get('data_period', 'N/A')),
        ]
        
        for stat_name, stat_value in ml_stats:
            worksheet.cell(row=row, column=1, value=stat_name)
            worksheet.cell(row=row, column=2, value=stat_value)
            row += 1
    
    row += 1
    if 'data_quality' in ml_analysis:
        worksheet.cell(row=row, column=1, value="Kualitas Data")
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        quality = ml_analysis['data_quality']
        quality_stats = [
            ("Total Tahun", quality.get('total_years', 0)),
            ("Total Records", quality.get('total_records', 0)),
            ("Kelengkapan", quality.get('completeness', 'N/A')),
        ]
        
        for stat_name, stat_value in quality_stats:
            worksheet.cell(row=row, column=1, value=stat_name)
            worksheet.cell(row=row, column=2, value=stat_value)
            row += 1

# ===== HELPER FUNCTIONS =====
def calculate_guest_count(occupied_rooms):
    """
    Calculate random guest count for occupied rooms
    Probability: 1 person=20%, 2 persons=40%, 3 persons=40%, 4 persons=10%
    """
    total_guests = 0
    for _ in range(occupied_rooms):
        rand = random.random()
        if rand < 0.20:
            guests = 1
        elif rand < 0.60:  # 20% + 40% = 60%
            guests = 2
        elif rand < 0.90:  # 60% + 30% = 90% (adjusted to make 3 persons 30%)
            guests = 3
        else:
            guests = 4
        total_guests += guests
    return total_guests

def calculate_children(male, female):
    """
    Calculate random children count (10-20% of adults)
    Returns (male_child, female_child)
    """
    # Random percentage between 10-20%
    percentage = random.uniform(0.10, 0.20)
    
    male_child = int(male * percentage)
    female_child = int(female * percentage)
    
    return male_child, female_child

# ===== AUTHENTICATION ROUTES =====
@app.route('/')
def homepage():
    """Homepage - redirect based on authentication status"""
    if current_user.is_authenticated:
        # Redirect based on role
        if current_user.role == 'admin':
            return redirect(url_for('admin_home'))
        elif current_user.role == 'hotel':
            return redirect(url_for('hotel_home'))
        elif current_user.role == 'tourism':
            return redirect(url_for('tourism_home'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        # Already logged in, redirect to appropriate home
        if current_user.role == 'admin':
            return redirect(url_for('admin_home'))
        elif current_user.role == 'hotel':
            return redirect(url_for('hotel_home'))
        elif current_user.role == 'tourism':
            return redirect(url_for('tourism_home'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('Username dan password harus diisi', 'error')
            return redirect(url_for('login'))
        
        user = User.get_by_username(app.config['DATABASE'], username)
        
        if user and user.check_password(password):
            login_user(user)
            flash(f'Selamat datang, {user.username}!', 'success')
            
            # Redirect based on role
            if user.role == 'admin':
                return redirect(url_for('admin_home'))
            elif user.role == 'hotel':
                return redirect(url_for('hotel_home'))
            elif user.role == 'tourism':
                return redirect(url_for('tourism_home'))
        else:
            flash('Username atau password salah', 'error')
            return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Logout user"""
    logout_user()
    flash('Anda telah logout', 'success')
    return redirect(url_for('login'))

# ===== ADMIN ROUTES =====
@app.route('/admin/home')
@login_required
@role_required('admin')
def admin_home():
    """Admin homepage"""
    # Get statistics
    conn = get_db_connection()
    
    cursor = conn.execute('SELECT COUNT(*) as count FROM users')
    total_users = cursor.fetchone()['count']
    
    cursor = conn.execute('SELECT COUNT(*) as count FROM hotel_data')
    total_hotel_data = cursor.fetchone()['count']
    
    cursor = conn.execute('SELECT COUNT(*) as count FROM tourism_site_data')
    total_tourism_data = cursor.fetchone()['count']
    
    cursor = conn.execute('SELECT COUNT(*) as count FROM tourism_data')
    total_ml_data = cursor.fetchone()['count']
    
    conn.close()
    
    return render_template('admin/home.html',
                         total_users=total_users,
                         total_hotel_data=total_hotel_data,
                         total_tourism_data=total_tourism_data,
                         total_ml_data=total_ml_data)

@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    """Admin user management - list all users"""
    users = User.get_all(app.config['DATABASE'])
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_create_user():
    """Admin create new user"""
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        
        if not all([username, email, password, role]):
            flash('Semua field harus diisi', 'error')
            return redirect(url_for('admin_create_user'))
        
        # Validate role
        if role not in ['hotel', 'tourism']:
            flash('Role harus hotel atau tourism', 'error')
            return redirect(url_for('admin_create_user'))
        
        # Check if username exists
        existing_user = User.get_by_username(app.config['DATABASE'], username)
        if existing_user:
            flash('Username sudah digunakan', 'error')
            return redirect(url_for('admin_create_user'))
        
        # Create user
        try:
            User.create(app.config['DATABASE'], username, password, role, email)
            flash(f'User {username} berhasil dibuat dengan role {role}!', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            flash(f'Error membuat user: {str(e)}', 'error')
            return redirect(url_for('admin_create_user'))
    
    return render_template('admin/create_user.html')

@app.route('/admin/hotel-data')
@login_required
@role_required('admin')
def admin_hotel_data():
    """Admin view all hotel data from all users"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all hotel data with user and hotel info
    cursor.execute('''
        SELECT 
            hd.id,
            hd.date,
            hd.occupied_rooms,
            hd.guest_count,
            hi.hotel_name,
            hi.total_rooms,
            u.username,
            hd.created_at
        FROM hotel_data hd
        JOIN hotel_info hi ON hd.user_id = hi.user_id
        JOIN users u ON hd.user_id = u.id
        ORDER BY hd.date DESC
    ''')
    
    rows = cursor.fetchall()
    conn.close()
    
    data = []
    for row in rows:
        occupancy_rate = (row['occupied_rooms'] / row['total_rooms'] * 100) if row['total_rooms'] > 0 else 0
        data.append({
            'id': row['id'],
            'date': row['date'],
            'hotel_name': row['hotel_name'],
            'username': row['username'],
            'occupied_rooms': row['occupied_rooms'],
            'total_rooms': row['total_rooms'],
            'guest_count': row['guest_count'],
            'occupancy_rate': round(occupancy_rate, 1),
            'created_at': row['created_at']
        })
    
    return render_template('admin/hotel_data.html', data=data)

@app.route('/admin/tourism-data')
@login_required
@role_required('admin')
def admin_tourism_data():
    """Admin view all tourism data from all users"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all tourism data with user info
    cursor.execute('''
        SELECT 
            td.id,
            td.date,
            td.origin,
            td.total_visitors,
            td.male_adult,
            td.female_adult,
            td.male_child,
            td.female_child,
            u.username,
            td.created_at
        FROM tourism_site_data td
        JOIN users u ON td.user_id = u.id
        ORDER BY td.date DESC
    ''')
    
    rows = cursor.fetchall()
    conn.close()
    
    data = []
    for row in rows:
        total_adults = row['male_adult'] + row['female_adult']
        total_children = row['male_child'] + row['female_child']
        data.append({
            'id': row['id'],
            'date': row['date'],
            'origin': row['origin'],
            'username': row['username'],
            'total_visitors': row['total_visitors'],
            'male_adult': row['male_adult'],
            'female_adult': row['female_adult'],
            'male_child': row['male_child'],
            'female_child': row['female_child'],
            'total_adults': total_adults,
            'total_children': total_children,
            'created_at': row['created_at']
        })
    
    return render_template('admin/tourism_data.html', data=data)

@app.route('/admin/hotel-data/export/<format>')
@login_required
@role_required('admin')
def admin_export_hotel_data(format):
    """Admin export all hotel data"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all hotel data
    cursor.execute('''
        SELECT 
            hd.date,
            hi.hotel_name,
            u.username,
            hd.occupied_rooms,
            hi.total_rooms,
            hd.guest_count,
            hd.created_at
        FROM hotel_data hd
        JOIN hotel_info hi ON hd.user_id = hi.user_id
        JOIN users u ON hd.user_id = u.id
        ORDER BY hd.date DESC
    ''')
    
    rows = cursor.fetchall()
    conn.close()
    
    # Prepare data
    data = []
    for row in rows:
        occupancy_rate = (row['occupied_rooms'] / row['total_rooms'] * 100) if row['total_rooms'] > 0 else 0
        data.append({
            'Tanggal': row['date'],
            'Hotel': row['hotel_name'],
            'Petugas': row['username'],
            'Kamar Terisi': row['occupied_rooms'],
            'Total Kamar': row['total_rooms'],
            'Tingkat Okupansi (%)': round(occupancy_rate, 1),
            'Jumlah Tamu': row['guest_count']
        })
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"admin_hotel_data_{timestamp}"
    
    # Create DataFrame for export
    df = pd.DataFrame(data)
    
    # Export based on format
    if format == 'excel':
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data Hotel')
        output.seek(0)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename += '.xlsx'
    elif format == 'csv':
        output = io.BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        mimetype = 'text/csv; charset=utf-8'
        filename += '.csv'
    elif format == 'pdf':
        # For PDF, use ReportLab
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
            alignment=1
        )
        
        title = Paragraph("Laporan Data Hotel - Admin", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        info_text = f"Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [list(df.columns)]
        for _, row in df.iterrows():
            table_data.append(list(row))
        
        table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        mimetype = 'application/pdf'
        filename += '.pdf'
    else:
        flash('Format tidak valid', 'error')
        return redirect(url_for('admin_hotel_data'))
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype
    )

@app.route('/admin/tourism-data/export/<format>')
@login_required
@role_required('admin')
def admin_export_tourism_data(format):
    """Admin export all tourism data"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all tourism data
    cursor.execute('''
        SELECT 
            td.date,
            td.origin,
            u.username,
            td.total_visitors,
            td.male_adult,
            td.female_adult,
            td.male_child,
            td.female_child,
            td.created_at
        FROM tourism_site_data td
        JOIN users u ON td.user_id = u.id
        ORDER BY td.date DESC
    ''')
    
    rows = cursor.fetchall()
    conn.close()
    
    # Prepare data
    data = []
    for row in rows:
        data.append({
            'Tanggal': row['date'],
            'Asal': row['origin'],
            'Petugas': row['username'],
            'Total Pengunjung': row['total_visitors'],
            'Dewasa Laki-laki': row['male_adult'],
            'Dewasa Perempuan': row['female_adult'],
            'Anak Laki-laki': row['male_child'],
            'Anak Perempuan': row['female_child']
        })
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"admin_tourism_data_{timestamp}"
    
    # Create DataFrame for export
    df = pd.DataFrame(data)
    
    # Export based on format
    if format == 'excel':
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data Wisata')
        output.seek(0)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename += '.xlsx'
    elif format == 'csv':
        output = io.BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        mimetype = 'text/csv; charset=utf-8'
        filename += '.csv'
    elif format == 'pdf':
        # For PDF, use ReportLab
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
            alignment=1
        )
        
        title = Paragraph("Laporan Data Wisata - Admin", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        info_text = f"Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [list(df.columns)]
        for _, row in df.iterrows():
            table_data.append(list(row))
        
        table = Table(table_data, colWidths=[1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.9*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        mimetype = 'application/pdf'
        filename += '.pdf'
    else:
        flash('Format tidak valid', 'error')
        return redirect(url_for('admin_tourism_data'))
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype
    )

@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_user(user_id):
    """Admin edit user"""
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        role = request.form.get('role')
        password = request.form.get('password')
        
        # Get current user to check if username changed
        current_user_data = User.get_by_id(app.config['DATABASE'], user_id)
        if not current_user_data:
            flash('User tidak ditemukan', 'error')
            return redirect(url_for('admin_users'))
        
        # Validate inputs
        if not username or not email or not role:
            flash('Username, email, dan role harus diisi', 'error')
            return redirect(url_for('admin_edit_user', user_id=user_id))
        
        # Validate role
        if role not in ['admin', 'hotel', 'tourism']:
            flash('Role tidak valid', 'error')
            return redirect(url_for('admin_edit_user', user_id=user_id))
        
        # Check if username changed and if new username exists
        if username != current_user_data.username:
            existing_user = User.get_by_username(app.config['DATABASE'], username)
            if existing_user:
                flash('Username sudah digunakan', 'error')
                return redirect(url_for('admin_edit_user', user_id=user_id))
        
        # Update user
        try:
            User.update(
                app.config['DATABASE'],
                user_id,
                username=username,
                email=email,
                role=role,
                password=password if password else None
            )
            flash(f'User {username} berhasil diupdate!', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            flash(f'Error mengupdate user: {str(e)}', 'error')
            return redirect(url_for('admin_edit_user', user_id=user_id))
    
    # GET request - show edit form
    user = User.get_by_id(app.config['DATABASE'], user_id)
    if not user:
        flash('User tidak ditemukan', 'error')
        return redirect(url_for('admin_users'))
    
    return render_template('admin/edit_user.html', user=user)



# ===== EXISTING ROUTES (NOW PROTECTED FOR ADMIN ONLY) =====
@app.route('/upload', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def upload():
    if request.method == 'POST':
        if 'csv_file' not in request.files:
            flash('Tidak ada file yang dipilih', 'error')
            return redirect(request.url)
        
        file = request.files['csv_file']
        year = request.form.get('year')
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'error')
            return redirect(request.url)
        
        if not year:
            flash('Tahun harus diisi', 'error')
            return redirect(request.url)
        
        if file.filename and not file.filename.lower().endswith('.csv'):
            flash('File harus berformat CSV', 'error')
            return redirect(request.url)
        
        if not validate_year(year):
            flash('Tahun harus antara 2000 dan tahun depan', 'error')
            return redirect(request.url)
        
        try:
            year = int(year)
        except ValueError:
            flash('Tahun harus berupa angka', 'error')
            return redirect(request.url)
        
        filename = f"tourism_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            df = pd.read_csv(filepath)
            palembang_found = False
            for index, row in df.iterrows():
                for cell in row:
                    if 'Palembang' in str(cell):
                        palembang_found = True
                        break
                if palembang_found:
                    break
            
            if not palembang_found:
                os.remove(filepath)
                flash('Data Palembang tidak ditemukan dalam file', 'error')
                return redirect(request.url)
                
        except Exception as e:
            os.remove(filepath)
            flash(f'File tidak bisa dibaca: {str(e)}', 'error')
            return redirect(request.url)
        
        success, message = process_csv_file_simple(filepath, year)
        
        if success:
            conn = get_db_connection()
            conn.execute(
                'INSERT INTO uploaded_files (filename, year) VALUES (?, ?)',
                (filename, year)
            )
            conn.commit()
            conn.close()
            flash(f'File berhasil diupload: {message}', 'success')
        else:
            flash(f'Error: {message}', 'error')
            if os.path.exists(filepath):
                os.remove(filepath)
        
        return redirect(url_for('upload'))
    
    db_stats = data_processor.get_database_stats()
    uploaded_files = data_processor.get_uploaded_files_info()
    
    return render_template('upload.html', 
                         db_stats=db_stats, 
                         uploaded_files=uploaded_files)

@app.route('/upload-pdf', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def upload_pdf():
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            flash('Tidak ada file PDF yang dipilih', 'error')
            return redirect(request.url)
        
        file = request.files['pdf_file']
        year = request.form.get('year')
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'error')
            return redirect(request.url)
        
        if file.filename and not file.filename.lower().endswith('.pdf'):
            flash('File harus berformat PDF', 'error')
            return redirect(request.url)
        
        if year and not validate_year(year):
            flash('Tahun harus antara 2000 dan tahun depan', 'error')
            return redirect(request.url)
        
        try:
            year_int = int(year) if year else None
        except ValueError:
            flash('Tahun harus berupa angka', 'error')
            return redirect(request.url)
        
        filename = f"tourism_pdf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            success, message = pdf_processor.process_pdf_for_database(filepath, year_int)
            
            if success:
                conn = get_db_connection()
                conn.execute(
                    'INSERT INTO uploaded_files (filename, year) VALUES (?, ?)',
                    (filename, year_int if year_int else datetime.now().year)
                )
                conn.commit()
                conn.close()
                
                flash(f'PDF berhasil diproses: {message}', 'success')
            else:
                flash(f'Error processing PDF: {message}', 'error')
            
            if os.path.exists(filepath):
                os.remove(filepath)
                
        except Exception as e:
            flash(f'Error processing PDF: {str(e)}', 'error')
            if os.path.exists(filepath):
                os.remove(filepath)
        
        return redirect(url_for('upload_pdf'))
    
    db_stats = data_processor.get_database_stats()
    uploaded_files = data_processor.get_uploaded_files_info()
    current_year = datetime.now().year  # TAMBAHKAN INI
    
    return render_template('upload_pdf.html', 
                         db_stats=db_stats, 
                         uploaded_files=uploaded_files,
                         current_year=current_year)  # TAMBAHKAN INI

@app.route('/convert-pdf-to-csv', methods=['POST'])
def convert_pdf_to_csv():
    if 'pdf_file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'})
    
    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'message': 'File must be PDF'})
    
    temp_pdf = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    file.save(temp_pdf)
    
    try:
        output_csv = temp_pdf.replace('.pdf', '.csv')
        success, message = pdf_processor.pdf_to_csv(temp_pdf, output_csv)
        
        if success:
            return send_file(
                output_csv,
                as_attachment=True,
                download_name=f"converted_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mimetype='text/csv'
            )
        else:
            return jsonify({'success': False, 'message': message})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    
    finally:
        if os.path.exists(temp_pdf):
            os.remove(temp_pdf)
        if os.path.exists(output_csv):
            os.remove(output_csv)

@app.route('/dashboard')
@login_required
@role_required('admin')
def dashboard():
    conn = get_db_connection()
    query = '''
        SELECT year, month, value 
        FROM tourism_data 
        ORDER BY year, 
        CASE month
            WHEN 'January' THEN 1
            WHEN 'February' THEN 2
            WHEN 'March' THEN 3
            WHEN 'April' THEN 4
            WHEN 'May' THEN 5
            WHEN 'June' THEN 6
            WHEN 'July' THEN 7
            WHEN 'August' THEN 8
            WHEN 'September' THEN 9
            WHEN 'October' THEN 10
            WHEN 'November' THEN 11
            WHEN 'December' THEN 12
        END
    '''
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    analysis_results = analyze_data()
    
    try:
        ml_analysis = ml_analyzer.get_detailed_analysis()
        ml_analysis['suggestions'] = ml_analysis['suggestions'][:3]
    except Exception as e:
        ml_analysis = {
            'suggestions': ["ML Analysis sedang dalam perbaikan"],
            'patterns': {},
            'summary': {},
            'data_quality': {'total_years': 0, 'total_records': 0}
        }
    
    charts_data_advanced = {}
    
    db_stats = data_processor.get_database_stats()
    data_complexity = get_data_complexity_level()
    
    return render_template('dashboard.html',
                         **analysis_results,
                         ml_analysis=ml_analysis,
                         charts_data_advanced=charts_data_advanced,
                         db_stats=db_stats,
                         data_complexity=data_complexity,
                         df_empty=df.empty)

@app.route('/delete-data', methods=['POST'])
def delete_data():
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM tourism_data')
        conn.execute('DELETE FROM uploaded_files')
        conn.commit()
        conn.close()
        
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if os.path.isfile(file_path):
                os.unlink(file_path)
        
        flash('Semua data berhasil dihapus', 'success')
    except Exception as e:
        flash(f'Error menghapus data: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/export-excel')
def export_excel():
    try:
        conn = get_db_connection()
        query = '''
            SELECT year, month, value 
            FROM tourism_data 
            ORDER BY year, 
            CASE month
                WHEN 'January' THEN 1
                WHEN 'February' THEN 2
                WHEN 'March' THEN 3
                WHEN 'April' THEN 4
                WHEN 'May' THEN 5
                WHEN 'June' THEN 6
                WHEN 'July' THEN 7
                WHEN 'August' THEN 8
                WHEN 'September' THEN 9
                WHEN 'October' THEN 10
                WHEN 'November' THEN 11
                WHEN 'December' THEN 12
            END
        '''
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        ml_analysis = ml_analyzer.get_detailed_analysis()
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        ws_raw = wb.create_sheet("Data Mentah")
        _create_raw_data_sheet(ws_raw, df)
        
        ws_ml = wb.create_sheet("Analisis ML")
        _create_ml_analysis_sheet(ws_ml, ml_analysis)
        
        ws_charts = wb.create_sheet("Visualisasi")
        _create_charts_sheet(ws_charts, df)
        
        ws_stats = wb.create_sheet("Statistik")
        _create_statistics_sheet(ws_stats, df, ml_analysis)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"tourism_analysis_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            attachment_filename=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating Excel file: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/api/chart-data')
def chart_data():
    analysis_results = analyze_data()
    return jsonify(analysis_results['charts_data'])

@app.route('/api/advanced-chart-data')
def advanced_chart_data():
    conn = get_db_connection()
    query = 'SELECT year, month, value FROM tourism_data'
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    try:
        charts_data = chart_generator.generate_all_charts_data(df)
        return jsonify(charts_data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/analysis-data')
def analysis_data():
    try:
        analysis_results = ml_analyzer.get_detailed_analysis()
        analysis_results['suggestions'] = analysis_results['suggestions'][:3]
        return jsonify(analysis_results)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/db-stats')
def db_stats_api():
    stats = data_processor.get_database_stats()
    return jsonify(stats)

@app.errorhandler(413)
def too_large(e):
    flash('File terlalu besar. Maksimal 16MB', 'error')
    return redirect(url_for('upload'))

@app.errorhandler(500)
def internal_error(error):
    flash('Terjadi error internal server. Silakan coba lagi.', 'error')
    return redirect(url_for('homepage'))

def not_found(error):
    return render_template('404.html'), 404

# ===== HOTEL STAFF ROUTES =====
@app.route('/hotel/home')
@login_required
@role_required('hotel')
def hotel_home():
    """Hotel staff homepage"""
    hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
    return render_template('hotel/home.html', hotel_info=hotel_info)

@app.route('/hotel/setup', methods=['GET', 'POST'])
@login_required
@role_required('hotel')
def hotel_setup():
    """Hotel setup - first time configuration"""
    # Check if already setup
    hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
    if hotel_info:
        flash('Hotel sudah dikonfigurasi. Hubungi admin untuk mengubah.', 'warning')
        return redirect(url_for('hotel_home'))
    
    if request.method == 'POST':
        hotel_name = request.form.get('hotel_name')
        total_rooms = request.form.get('total_rooms')
        
        if not hotel_name or not total_rooms:
            flash('Semua field harus diisi', 'error')
            return redirect(url_for('hotel_setup'))
        
        try:
            total_rooms = int(total_rooms)
            if total_rooms <= 0:
                flash('Jumlah kamar harus lebih dari 0', 'error')
                return redirect(url_for('hotel_setup'))
        except ValueError:
            flash('Jumlah kamar harus berupa angka', 'error')
            return redirect(url_for('hotel_setup'))
        
        HotelData.set_hotel_info(app.config['DATABASE'], current_user.id, hotel_name, total_rooms)
        flash(f'Hotel {hotel_name} berhasil dikonfigurasi!', 'success')
        return redirect(url_for('hotel_home'))
    
    return render_template('hotel/setup.html')

@app.route('/hotel/input', methods=['GET', 'POST'])
@login_required
@role_required('hotel')
def hotel_input():
    """Hotel daily data input"""
    hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
    
    if not hotel_info:
        flash('Silakan setup hotel terlebih dahulu', 'warning')
        return redirect(url_for('hotel_setup'))
    
    if request.method == 'POST':
        date = request.form.get('date')
        occupied_rooms = request.form.get('occupied_rooms')
        
        if not date or not occupied_rooms:
            flash('Semua field harus diisi', 'error')
            return redirect(url_for('hotel_input'))
        
        try:
            occupied_rooms = int(occupied_rooms)
            if occupied_rooms < 0:
                flash('Jumlah kamar terisi tidak boleh negatif', 'error')
                return redirect(url_for('hotel_input'))
            if occupied_rooms > hotel_info['total_rooms']:
                flash(f'Jumlah kamar terisi tidak boleh lebih dari total kamar ({hotel_info["total_rooms"]})', 'error')
                return redirect(url_for('hotel_input'))
        except ValueError:
            flash('Jumlah kamar terisi harus berupa angka', 'error')
            return redirect(url_for('hotel_input'))
        
        # Check if date already exists
        if HotelData.check_date_exists(app.config['DATABASE'], current_user.id, date):
            flash(f'Data untuk tanggal {date} sudah ada. Silakan edit di dashboard.', 'warning')
            return redirect(url_for('hotel_input'))
        
        # Calculate guest count
        guest_count = calculate_guest_count(occupied_rooms)
        
        # Save to database
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO hotel_data (user_id, date, occupied_rooms, guest_count)
            VALUES (?, ?, ?, ?)
        ''', (current_user.id, date, occupied_rooms, guest_count))
        conn.commit()
        conn.close()
        
        flash(f'Data berhasil disimpan! Jumlah tamu: {guest_count} orang', 'success')
        return redirect(url_for('hotel_dashboard'))
    
    return render_template('hotel/input.html', hotel_info=hotel_info)

@app.route('/hotel/dashboard')
@login_required
@role_required('hotel')
def hotel_dashboard():
    """Hotel dashboard with data table"""
    hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
    
    if not hotel_info:
        flash('Silakan setup hotel terlebih dahulu', 'warning')
        return redirect(url_for('hotel_setup'))
    
    data = HotelData.get_all_data(app.config['DATABASE'], current_user.id)
    
    return render_template('hotel/dashboard.html', 
                         hotel_info=hotel_info,
                         data=data)

@app.route('/hotel/edit/<int:data_id>', methods=['POST'])
@login_required
@role_required('hotel')
def hotel_edit(data_id):
    """Edit hotel data - only occupied_rooms"""
    occupied_rooms = request.form.get('occupied_rooms')
    
    if not occupied_rooms:
        flash('Jumlah kamar terisi harus diisi', 'error')
        return redirect(url_for('hotel_dashboard'))
    
    try:
        occupied_rooms = int(occupied_rooms)
        hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
        
        if occupied_rooms < 0:
            flash('Jumlah kamar terisi tidak boleh negatif', 'error')
            return redirect(url_for('hotel_dashboard'))
        
        if occupied_rooms > hotel_info['total_rooms']:
            flash(f'Jumlah kamar terisi tidak boleh lebih dari total kamar ({hotel_info["total_rooms"]})', 'error')
            return redirect(url_for('hotel_dashboard'))
        
        # Recalculate guest count
        guest_count = calculate_guest_count(occupied_rooms)
        
        # Update database
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE hotel_data 
            SET occupied_rooms = ?, guest_count = ?
            WHERE id = ? AND user_id = ?
        ''', (occupied_rooms, guest_count, data_id, current_user.id))
        conn.commit()
        conn.close()
        
        flash('Data berhasil diupdate!', 'success')
    except ValueError:
        flash('Jumlah kamar terisi harus berupa angka', 'error')
    
    return redirect(url_for('hotel_dashboard'))

@app.route('/hotel/export/<format>')
@login_required
@role_required('hotel')
def hotel_export(format):
    """Export hotel data to Excel, CSV, or PDF"""
    hotel_info = HotelData.get_hotel_info(app.config['DATABASE'], current_user.id)
    if not hotel_info:
        flash('Silakan setup hotel terlebih dahulu', 'warning')
        return redirect(url_for('hotel_setup'))
    
    data = HotelData.get_all_data(app.config['DATABASE'], current_user.id)
    
    if not data:
        flash('Tidak ada data untuk diexport', 'warning')
        return redirect(url_for('hotel_dashboard'))
    
    try:
        if format == 'excel':
            output = export_hotel_to_excel(data, hotel_info['hotel_name'], hotel_info['total_rooms'])
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"hotel_{hotel_info['hotel_name']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        elif format == 'csv':
            output = export_hotel_to_csv(data, hotel_info['hotel_name'], hotel_info['total_rooms'])
            mimetype = 'text/csv'
            filename = f"hotel_{hotel_info['hotel_name']}_{datetime.now().strftime('%Y%m%d')}.csv"
        elif format == 'pdf':
            output = export_hotel_to_pdf(data, hotel_info['hotel_name'], hotel_info['total_rooms'])
            mimetype = 'application/pdf'
            filename = f"hotel_{hotel_info['hotel_name']}_{datetime.now().strftime('%Y%m%d')}.pdf"
        else:
            flash('Format tidak valid', 'error')
            return redirect(url_for('hotel_dashboard'))
        
        if output:
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype=mimetype
            )
        else:
            flash('Gagal membuat file export', 'error')
            return redirect(url_for('hotel_dashboard'))
    except Exception as e:
        flash(f'Error export: {str(e)}', 'error')
        return redirect(url_for('hotel_dashboard'))

# ===== TOURISM STAFF ROUTES =====
@app.route('/tourism/home')
@login_required
@role_required('tourism')
def tourism_home():
    """Tourism staff homepage"""
    return render_template('tourism/home.html')

@app.route('/tourism/input', methods=['GET', 'POST'])
@login_required
@role_required('tourism')
def tourism_input():
    """Tourism daily data input"""
    if request.method == 'POST':
        date = request.form.get('date')
        origin = request.form.get('origin')
        total_visitors = request.form.get('total_visitors')
        male = request.form.get('male')
        female = request.form.get('female')
        male_child_input = request.form.get('male_child', '').strip()
        female_child_input = request.form.get('female_child', '').strip()
        
        if not all([date, origin, total_visitors, male, female]):
            flash('Field tanggal, asal, total, laki-laki, dan perempuan harus diisi', 'error')
            return redirect(url_for('tourism_input'))
        
        try:
            total_visitors = int(total_visitors)
            male = int(male)
            female = int(female)
            
            if total_visitors < 0 or male < 0 or female < 0:
                flash('Jumlah pengunjung tidak boleh negatif', 'error')
                return redirect(url_for('tourism_input'))
            
            # Calculate children if not provided
            if male_child_input and female_child_input:
                male_child = int(male_child_input)
                female_child = int(female_child_input)
            else:
                male_child, female_child = calculate_children(male, female)
            
            # Validate total
            total_calculated = male + female
            if total_calculated > total_visitors:
                flash('Jumlah laki-laki + perempuan tidak boleh lebih dari total pengunjung', 'error')
                return redirect(url_for('tourism_input'))
            
        except ValueError:
            flash('Semua field angka harus berupa angka valid', 'error')
            return redirect(url_for('tourism_input'))
        
        # Check if date already exists
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id FROM tourism_site_data 
            WHERE user_id = ? AND date = ?
        ''', (current_user.id, date))
        
        if cursor.fetchone():
            conn.close()
            flash(f'Data untuk tanggal {date} sudah ada. Silakan edit di dashboard.', 'warning')
            return redirect(url_for('tourism_input'))
        
        # Calculate adults (subtract children)
        male_adult = male - male_child
        female_adult = female - female_child
        
        # Save to database
        cursor.execute('''
            INSERT INTO tourism_site_data 
            (user_id, date, origin, total_visitors, male_adult, female_adult, male_child, female_child)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (current_user.id, date, origin, total_visitors, male_adult, female_adult, male_child, female_child))
        conn.commit()
        conn.close()
        
        flash('Data berhasil disimpan!', 'success')
        return redirect(url_for('tourism_dashboard'))
    
    return render_template('tourism/input.html')

@app.route('/tourism/dashboard')
@login_required
@role_required('tourism')
def tourism_dashboard():
    """Tourism dashboard with data table"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM tourism_site_data 
        WHERE user_id = ?
        ORDER BY date DESC
    ''', (current_user.id,))
    
    rows = cursor.fetchall()
    conn.close()
    
    data = []
    for row in rows:
        data.append({
            'id': row['id'],
            'date': row['date'],
            'origin': row['origin'],
            'total_visitors': row['total_visitors'],
            'male_adult': row['male_adult'],
            'female_adult': row['female_adult'],
            'male_child': row['male_child'],
            'female_child': row['female_child'],
            'created_at': row['created_at']
        })
    
    return render_template('tourism/dashboard.html', data=data)

@app.route('/tourism/edit/<int:data_id>', methods=['POST'])
@login_required
@role_required('tourism')
def tourism_edit(data_id):
    """Edit tourism data"""
    origin = request.form.get('origin')
    total_visitors = request.form.get('total_visitors')
    male_adult = request.form.get('male_adult')
    female_adult = request.form.get('female_adult')
    male_child = request.form.get('male_child')
    female_child = request.form.get('female_child')
    
    if not all([origin, total_visitors, male_adult, female_adult, male_child, female_child]):
        flash('Semua field harus diisi', 'error')
        return redirect(url_for('tourism_dashboard'))
    
    try:
        total_visitors = int(total_visitors)
        male_adult = int(male_adult)
        female_adult = int(female_adult)
        male_child = int(male_child)
        female_child = int(female_child)
        
        if any(x < 0 for x in [total_visitors, male_adult, female_adult, male_child, female_child]):
            flash('Jumlah pengunjung tidak boleh negatif', 'error')
            return redirect(url_for('tourism_dashboard'))
        
        # Update database
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE tourism_site_data 
            SET origin = ?, total_visitors = ?, male_adult = ?, female_adult = ?, male_child = ?, female_child = ?
            WHERE id = ? AND user_id = ?
        ''', (origin, total_visitors, male_adult, female_adult, male_child, female_child, data_id, current_user.id))
        conn.commit()
        conn.close()
        
        flash('Data berhasil diupdate!', 'success')
    except ValueError:
        flash('Semua field angka harus berupa angka valid', 'error')
    
    return redirect(url_for('tourism_dashboard'))

@app.route('/tourism/export/<format>')
@login_required
@role_required('tourism')
def tourism_export(format):
    """Export tourism data to Excel, CSV, or PDF"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM tourism_site_data 
        WHERE user_id = ?
        ORDER BY date DESC
    ''', (current_user.id,))
    
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        flash('Tidak ada data untuk diexport', 'warning')
        return redirect(url_for('tourism_dashboard'))
    
    data = []
    for row in rows:
        data.append({
            'id': row['id'],
            'date': row['date'],
            'origin': row['origin'],
            'total_visitors': row['total_visitors'],
            'male_adult': row['male_adult'],
            'female_adult': row['female_adult'],
            'male_child': row['male_child'],
            'female_child': row['female_child']
        })
    
    try:
        if format == 'excel':
            output = export_tourism_to_excel(data, current_user.username)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"tourism_{current_user.username}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        elif format == 'csv':
            output = export_tourism_to_csv(data, current_user.username)
            mimetype = 'text/csv'
            filename = f"tourism_{current_user.username}_{datetime.now().strftime('%Y%m%d')}.csv"
        elif format == 'pdf':
            output = export_tourism_to_pdf(data, current_user.username)
            mimetype = 'application/pdf'
            filename = f"tourism_{current_user.username}_{datetime.now().strftime('%Y%m%d')}.pdf"
        else:
            flash('Format tidak valid', 'error')
            return redirect(url_for('tourism_dashboard'))
        
        if output:
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype=mimetype
            )
        else:
            flash('Gagal membuat file export', 'error')
            return redirect(url_for('tourism_dashboard'))
    except Exception as e:
        flash(f'Error export: {str(e)}', 'error')
        return redirect(url_for('tourism_dashboard'))

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404


if __name__ == '__main__':
    init_db()
    create_default_admin()
    print("=== Tourism Data Management System ===")
    print("Server running on: http://127.0.0.1:5000")
    print("Routes available:")
    print("  /login         - Login page")
    print("  /admin/*       - Admin routes")
    print("  /hotel/*       - Hotel staff routes")
    print("  /tourism/*     - Tourism staff routes")
    print("  /api/*         - Various API endpoints")
    app.run(debug=True, host='0.0.0.0', port=5000, )